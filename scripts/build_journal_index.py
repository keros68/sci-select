"""Build a user-owned sci-select journal index from local data files.

This module is intentionally data-source agnostic. It ships import logic and a
stable JSON schema, but it does not bundle third-party journal datasets.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


INDEX_SCHEMA = "sci-select-journal-index-v1"
JCR_RELEASE_YEAR = 2026
JCR_DATA_YEAR = 2025


def build_index(
    cas_2025_xlsx: str = "",
    xinrui_2026_xlsx: str = "",
    jcr_file: str = "",
    showjcr_db: str = "",
    showjcr_tables: Optional[Sequence[str]] = None,
) -> Dict:
    """Build a JSON-serializable journal index payload from local user files."""
    merged: Dict[str, Dict] = {}
    source_types: List[str] = []

    if cas_2025_xlsx:
        _merge_rows(merged, read_cas_2025_file(cas_2025_xlsx))
        source_types.append("cas_2025")

    if xinrui_2026_xlsx:
        _merge_rows(merged, read_xinrui_2026_file(xinrui_2026_xlsx))
        source_types.append("xinrui_2026")

    if jcr_file:
        _merge_rows(merged, read_jcr_2025_file(jcr_file))
        source_types.append("jcr_2025")

    if showjcr_db:
        _merge_rows(merged, read_showjcr_db(showjcr_db, showjcr_tables))
        source_types.append("showjcr_db")

    rows = _unique_rows(merged.values())
    rows = sorted(rows, key=lambda row: _normalize_name(row.get("title", "")))
    return {
        "meta": {
            "schema": INDEX_SCHEMA,
            "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
            "source_types": source_types,
            "jcr_release_year": JCR_RELEASE_YEAR,
            "jcr_data_year": JCR_DATA_YEAR,
            "note": "Generated from user-supplied local files. Third-party datasets are not bundled with sci-select.",
        },
        "journals": rows,
    }


def write_index(payload: Dict, output_path: str) -> None:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_sqlite_index(payload: Dict, output_path: str) -> None:
    """Write payload into sci-select's own SQLite schema."""
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        output.unlink()

    conn = sqlite3.connect(output)
    try:
        conn.execute(
            """
            CREATE TABLE metadata (
                key TEXT PRIMARY KEY,
                value TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE journals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                normalized_title TEXT NOT NULL,
                issn TEXT,
                normalized_issn TEXT,
                eissn TEXT,
                normalized_eissn TEXT,
                payload_json TEXT NOT NULL
            )
            """
        )
        conn.execute("CREATE INDEX idx_journals_normalized_title ON journals(normalized_title)")
        conn.execute("CREATE INDEX idx_journals_normalized_issn ON journals(normalized_issn)")
        conn.execute("CREATE INDEX idx_journals_normalized_eissn ON journals(normalized_eissn)")

        meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
        meta = {
            **meta,
            "sqlite_schema": "sci-select-sqlite-v1",
            "journal_count": len(payload.get("journals", [])),
        }
        for key, value in meta.items():
            conn.execute(
                "INSERT INTO metadata(key, value) VALUES (?, ?)",
                (str(key), json.dumps(value, ensure_ascii=False) if not isinstance(value, str) else value),
            )

        for row in payload.get("journals", []):
            if not isinstance(row, dict) or not row.get("title"):
                continue
            conn.execute(
                """
                INSERT INTO journals(
                    title, normalized_title, issn, normalized_issn,
                    eissn, normalized_eissn, payload_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(row.get("title", "")),
                    _normalize_name(row.get("title", "")),
                    row.get("issn", ""),
                    _normalize_issn_for_lookup(row.get("issn", "")),
                    row.get("eissn", ""),
                    _normalize_issn_for_lookup(row.get("eissn", "")),
                    json.dumps(row, ensure_ascii=False, sort_keys=True),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def read_cas_2025_file(path: str) -> List[Dict]:
    rows = []
    for raw in _read_table_file(path):
        title = _pick(raw, "Journal", "期刊名称", "Title", "Journal name")
        if not title:
            continue
        partition = _normalize_partition(_pick(raw, "分区", "中科院分区", "cas_2025"))
        item = {"title": title}
        if partition:
            item["cas_2025"] = partition
            item.setdefault("tags", []).append(f"中科院{partition}")
        top = _normalize_yes_no(_pick(raw, "Top", "Top期刊", "是否Top"))
        if top:
            item["cas_top_2025"] = top == "是"
            if top == "是":
                item.setdefault("tags", []).append("中科院Top")
        oa = _normalize_yes_no(_pick(raw, "Open Access", "OA", "是否OA"))
        if oa:
            item["open_access"] = oa == "是"
        rows.append(item)
    return rows


def read_xinrui_2026_file(path: str) -> List[Dict]:
    rows = []
    for raw in _read_table_file(path):
        title = _pick(raw, "期刊名称", "Journal", "Journal name", "Title")
        partition = _normalize_partition(_pick(raw, "新锐分区", "新锐", "xuankan_2026", "XinRui"))
        if not title or not partition:
            continue
        issn1 = _normalize_issn(_pick(raw, "issn1", "ISSN", "Issn", "issn"))
        issn2 = _normalize_issn(_pick(raw, "issn2", "eISSN", "EISSN", "eissn"))
        item = {
            "title": title,
            "xuankan_2026": partition,
            "tags": [f"新锐{partition}"],
        }
        if issn1:
            item["issn"] = issn1
        if issn2 and issn2 != issn1:
            item["eissn"] = issn2
        subject = _pick(raw, "学科", "大类学科", "field", "Field")
        if subject:
            item["xinrui_subject"] = subject
        rows.append(item)
    return rows


def read_jcr_2025_file(path: str) -> List[Dict]:
    merged: Dict[str, Dict] = {}
    for raw in _read_table_file(path):
        item = _jcr_row_to_index(raw)
        if item:
            _merge_rows(merged, [item])
    return list(merged.values())


def read_showjcr_db(path: str, table_names: Optional[Sequence[str]] = None) -> List[Dict]:
    """Read compatible ShowJCR-style SQLite tables from a user-supplied jcr.db."""
    rows: List[Dict] = []
    conn = sqlite3.connect(path)
    try:
        conn.row_factory = sqlite3.Row
        tables = list(table_names) if table_names else _sqlite_tables(conn)
        for table in tables:
            columns = _sqlite_columns(conn, table)
            if "Journal" not in columns and "journal" not in {c.lower() for c in columns}:
                continue
            for raw in conn.execute(f'SELECT * FROM "{table}"'):
                row = {key: raw[key] for key in raw.keys()}
                mapped = _showjcr_row_to_index(table, row)
                if mapped:
                    rows.append(mapped)
    finally:
        conn.close()
    merged: Dict[str, Dict] = {}
    _merge_rows(merged, rows)
    return list(merged.values())


def _showjcr_row_to_index(table: str, row: Dict) -> Dict:
    table_upper = table.upper()
    if "XR2026" in table_upper:
        return _xinrui_like_row_to_index(row)
    if "FQBJCR2025" in table_upper or "CAS2025" in table_upper or "FENQU" in table_upper:
        return _cas_like_row_to_index(row)
    if "JCR2025" in table_upper:
        return _jcr_row_to_index(row)
    return _generic_row_to_index(row)


def _xinrui_like_row_to_index(row: Dict) -> Dict:
    title = _pick(row, "Journal", "期刊名称", "Journal name", "Title")
    partition = _normalize_partition(_pick(row, "新锐分区", "分区", "大类分区", "Tier"))
    if not title:
        return {}
    item = {"title": title}
    if partition:
        item["xuankan_2026"] = partition
        item.setdefault("tags", []).append(f"新锐{partition}")
    _copy_issns(row, item)
    return item


def _cas_like_row_to_index(row: Dict) -> Dict:
    title = _pick(row, "Journal", "期刊名称", "Journal name", "Title")
    partition = _normalize_partition(_pick(row, "分区", "大类分区", "中科院分区"))
    if not title:
        return {}
    item = {"title": title}
    if partition:
        item["cas_2025"] = partition
        item.setdefault("tags", []).append(f"中科院{partition}")
    _copy_issns(row, item)
    return item


def _jcr_row_to_index(row: Dict) -> Dict:
    title = _pick(row, "Journal name", "Journal Name", "Journal", "Title", "Full Journal Title")
    if not title:
        return {}
    item: Dict = {
        "title": title,
        "jcr_release_year": JCR_RELEASE_YEAR,
        "jcr_data_year": JCR_DATA_YEAR,
    }
    _copy_issns(row, item)
    jif = _pick(row, "JIF", "Journal Impact Factor", "Impact Factor", "IF", "IF(2025)")
    if jif:
        item["jif_2025"] = str(jif)
        item["if_year"] = str(JCR_DATA_YEAR)
    quartile = _normalize_quartile(_pick(row, "JIF Quartile", "JCR Quartile", "Quartile", "Q"))
    if quartile:
        item["jcr_quartile_2025"] = quartile
        item["jcr_quartile"] = quartile
        item.setdefault("tags", []).append(quartile)
    category = _pick(row, "JCR Category", "Category", "学科", "Web of Science Category")
    edition = _pick(row, "Edition", "JCR Edition", "收录类型")
    if edition:
        item.setdefault("tags", []).append(str(edition).upper().replace(" ", ""))
    if category or quartile or edition:
        category_item = {}
        if category:
            category_item["category"] = str(category)
        if quartile:
            category_item["quartile"] = quartile
        if edition:
            category_item["edition"] = str(edition)
        item["jcr_categories"] = [category_item]
    return item


def _generic_row_to_index(row: Dict) -> Dict:
    title = _pick(row, "Journal", "Journal name", "Title", "期刊名称")
    if not title:
        return {}
    item = {"title": title}
    _copy_issns(row, item)
    return item


def _copy_issns(source: Dict, target: Dict) -> None:
    issn = _normalize_issn(_pick(source, "ISSN", "issn", "issn1", "ISSN1"))
    eissn = _normalize_issn(_pick(source, "eISSN", "EISSN", "eissn", "issn2", "ISSN2"))
    if issn:
        target["issn"] = issn
    if eissn and eissn != issn:
        target["eissn"] = eissn


def _merge_rows(merged: Dict[str, Dict], rows: Iterable[Dict]) -> None:
    for row in rows:
        if not row or not row.get("title"):
            continue
        key = _row_key(row)
        name_key = f"name:{_normalize_name(row.get('title', ''))}"
        current_key = key
        if current_key not in merged and name_key in merged:
            current_key = name_key
        current = merged.setdefault(current_key, {"title": row["title"]})
        _merge_row(current, row)
        if current_key != key and key not in merged:
            merged[key] = current


def _merge_row(current: Dict, incoming: Dict) -> None:
    for key, value in incoming.items():
        if value in ("", None, [], {}):
            continue
        if key == "tags":
            tags = current.setdefault("tags", [])
            for tag in value:
                if tag and tag not in tags:
                    tags.append(tag)
        elif key == "jcr_categories":
            categories = current.setdefault("jcr_categories", [])
            for category in value:
                if category and category not in categories:
                    categories.append(category)
            current["jcr_quartile_2025"] = _best_quartile(
                [cat.get("quartile", "") for cat in categories] + [current.get("jcr_quartile_2025", "")]
            )
            current["jcr_quartile"] = current["jcr_quartile_2025"]
        elif key in {"jcr_quartile_2025", "jcr_quartile"} and current.get(key):
            current[key] = _best_quartile([current[key], value])
        elif key not in current or current.get(key) in ("", None, [], {}):
            current[key] = value


def _unique_rows(rows: Iterable[Dict]) -> List[Dict]:
    unique: List[Dict] = []
    seen = set()
    for row in rows:
        marker = id(row)
        if marker in seen:
            continue
        seen.add(marker)
        unique.append(row)
    return unique


def _row_key(row: Dict) -> str:
    issn = _normalize_issn(row.get("issn", ""))
    eissn = _normalize_issn(row.get("eissn", ""))
    if issn:
        return f"issn:{issn}"
    if eissn:
        return f"issn:{eissn}"
    return f"name:{_normalize_name(row.get('title', ''))}"


def _read_table_file(path: str) -> List[Dict]:
    suffix = Path(path).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        return _read_csv(path, delimiter)
    raise ValueError(f"Unsupported table file type: {path}")


def _read_xlsx(path: str) -> List[Dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading .xlsx files requires openpyxl. Install with `pip install openpyxl`.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        raw_rows = [
            [cell for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        ]
    finally:
        workbook.close()
    if not raw_rows:
        return []
    header_index = _find_header_index(raw_rows)
    headers = [_clean_header(value) for value in raw_rows[header_index]]
    rows = []
    for values in raw_rows[header_index + 1:]:
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        if row:
            rows.append(row)
    return rows


def _read_csv(path: str, delimiter: str) -> List[Dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        raw_rows = [row for row in reader if any(cell for cell in row)]
    if not raw_rows:
        return []
    header_index = _find_header_index(raw_rows)
    headers = [_clean_header(value) for value in raw_rows[header_index]]
    rows = []
    for values in raw_rows[header_index + 1:]:
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        if row:
            rows.append(row)
    return rows


def _find_header_index(rows: List[List]) -> int:
    for index, row in enumerate(rows[:10]):
        normalized = {_clean_header(value).lower() for value in row if _clean_header(value)}
        if (
            "journal" in normalized
            or "期刊名称" in normalized
            or "journal name" in normalized
            or "journal" in " ".join(normalized)
        ):
            return index
    return 0


def _sqlite_tables(conn: sqlite3.Connection) -> List[str]:
    rows = conn.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
    return [row[0] for row in rows]


def _sqlite_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    rows = conn.execute(f'PRAGMA table_info("{table}")').fetchall()
    return [row[1] for row in rows]


def _pick(row: Dict, *names: str):
    lower = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return str(value).strip()
        value = lower.get(str(name).strip().lower())
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _clean_header(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _normalize_issn(value: str) -> str:
    text = str(value or "").strip()
    if text in {"-", "—", "nan", "None"}:
        return ""
    return re.sub(r"[^0-9Xx-]", "", text).upper()


def _normalize_issn_for_lookup(value: str) -> str:
    return re.sub(r"[^0-9Xx]", "", str(value or "")).upper()


def _normalize_partition(value: str) -> str:
    match = re.search(r"([1-4])\s*区", str(value or ""))
    if match:
        return f"{match.group(1)}区"
    match = re.fullmatch(r"\s*([1-4])\s*", str(value or ""))
    if match:
        return f"{match.group(1)}区"
    return ""


def _normalize_quartile(value: str) -> str:
    match = re.search(r"Q\s*([1-4])", str(value or ""), re.IGNORECASE)
    return f"Q{match.group(1)}" if match else ""


def _best_quartile(values: Iterable[str]) -> str:
    quartiles = [_normalize_quartile(value) for value in values if value]
    if not quartiles:
        return ""
    return sorted(set(quartiles), key=lambda value: int(value[1]))[0]


def _normalize_yes_no(value: str) -> str:
    text = str(value or "").strip().lower()
    if text in {"是", "yes", "y", "true", "1"}:
        return "是"
    if text in {"否", "no", "n", "false", "0"}:
        return "否"
    return ""


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build a sci-select journal index from user-supplied local files.")
    parser.add_argument("--cas-2025-xlsx", default="", help="Path to a CAS 2025 partition .xlsx/.csv file.")
    parser.add_argument("--xinrui-2026-xlsx", default="", help="Path to a XinRui 2026 partition .xlsx/.csv file.")
    parser.add_argument("--jcr-file", default="", help="Path to a JCR 2025 .xlsx/.csv export.")
    parser.add_argument("--showjcr-db", default="", help="Path to a user-supplied ShowJCR-style jcr.db SQLite file.")
    parser.add_argument("--showjcr-table", action="append", default=[], help="Specific ShowJCR SQLite table to import. Repeatable.")
    parser.add_argument("--output", default="", help="Output JSON path for SCI_SELECT_JOURNAL_INDEX_PATH.")
    parser.add_argument("--sqlite-output", default="", help="Output SQLite path for SCI_SELECT_JOURNAL_INDEX_DB.")
    args = parser.parse_args(argv)
    if not args.output and not args.sqlite_output:
        parser.error("Provide --sqlite-output, --output, or both.")

    payload = build_index(
        cas_2025_xlsx=args.cas_2025_xlsx,
        xinrui_2026_xlsx=args.xinrui_2026_xlsx,
        jcr_file=args.jcr_file,
        showjcr_db=args.showjcr_db,
        showjcr_tables=args.showjcr_table or None,
    )
    if args.output:
        write_index(payload, args.output)
        print(f"Wrote {len(payload['journals'])} journals to {args.output}")
    if args.sqlite_output:
        write_sqlite_index(payload, args.sqlite_output)
        print(f"Wrote {len(payload['journals'])} journals to {args.sqlite_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
