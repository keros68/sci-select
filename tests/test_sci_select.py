import csv
import inspect
import importlib
import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import requests

from openpyxl import Workbook

from scripts.build_journal_index import build_index, write_index, write_sqlite_index
import scripts.journal_metrics as metrics
import scripts.letpub_client as letpub
selector = importlib.import_module("scripts.select_journals")
from scripts.select_journals import (
    infer_paper_profile,
    merge_paper_profile,
    rank_metric_records,
    format_selection_report,
    format_selection_matrix,
    interleave_candidate_groups,
    assign_candidate_labels,
    assign_submission_bands,
)
from scripts.similar_works import discover_journal_candidates
import scripts.similar_works as similar_works
import scripts.recommend as legacy_recommend
from scripts.official_finders import build_finder_checklist, format_finder_checklist
from scripts.scope_evidence import verify_official_scope, build_scope_verification_queue
from scripts.profile_consistency import compare_profiles, validate_profile
from scripts.benchmark_score import score_benchmark, auxiliary_exact_journal_metrics
from scripts.benchmark_run import run_benchmark
from scripts.benchmark_dataset import build_annotation_sheet
from scripts.audit_journal_index import audit_index
from scripts.journal_index_client import search_index_journals


class SciSelectTests(unittest.TestCase):
    def test_build_journal_index_merges_user_supplied_cas_and_xinrui_excels(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            cas_path = os.path.join(tmpdir, "cas_2025.xlsx")
            xinrui_path = os.path.join(tmpdir, "xinrui_2026.xlsx")
            output_path = os.path.join(tmpdir, "journal_index.json")

            cas_book = Workbook()
            cas_ws = cas_book.active
            cas_ws.append(["Journal", "分区", "Top", "Open Access"])
            cas_ws.append(["Environmental Pollution", "2", "否", "否"])
            cas_ws.append(["Journal of Cleaner Production", "1", "是", "否"])
            cas_book.save(cas_path)

            xinrui_book = Workbook()
            xinrui_ws = xinrui_book.active
            xinrui_ws.append(["2026年期刊分区表（新锐分区）", "", "", "", "", ""])
            xinrui_ws.append(["序号", "期刊名称", "学科", "新锐分区", "issn1", "issn2"])
            xinrui_ws.append([1, "ENVIRONMENTAL POLLUTION", "环境科学与生态学", "2 区", "0269-7491", "1873-6424"])
            xinrui_ws.append([2, "Journal of Cleaner Production", "环境科学与生态学", "1区", "0959-6526", "1879-1786"])
            xinrui_book.save(xinrui_path)

            payload = build_index(cas_2025_xlsx=cas_path, xinrui_2026_xlsx=xinrui_path)
            write_index(payload, output_path)

            rows = {row["title"].lower(): row for row in payload["journals"]}

            self.assertEqual(rows["environmental pollution"]["cas_2025"], "2区")
            self.assertEqual(rows["environmental pollution"]["xuankan_2026"], "2区")
            self.assertEqual(rows["environmental pollution"]["issn"], "0269-7491")
            self.assertEqual(rows["journal of cleaner production"]["cas_2025"], "1区")
            self.assertEqual(rows["journal of cleaner production"]["xuankan_2026"], "1区")
            self.assertIn("新锐1区", rows["journal of cleaner production"]["tags"])
            self.assertEqual(payload["meta"]["schema"], "sci-select-journal-index-v1")
            self.assertNotIn(tmpdir, json.dumps(payload, ensure_ascii=False))

            with open(output_path, "r", encoding="utf-8") as f:
                saved = json.load(f)
            self.assertEqual(len(saved["journals"]), 2)

    def test_build_journal_index_writes_own_sqlite_schema(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            sqlite_path = os.path.join(tmpdir, "sci_select_journals.sqlite")
            payload = {
                "meta": {"schema": "sci-select-journal-index-v1"},
                "journals": [
                    {
                        "title": "ENVIRONMENTAL POLLUTION",
                        "issn": "0269-7491",
                        "eissn": "1873-6424",
                        "jif_2025": "8.8",
                        "jcr_release_year": 2026,
                        "jcr_data_year": 2025,
                        "jcr_quartile_2025": "Q1",
                        "cas_2025": "2区",
                        "xuankan_2026": "2区",
                        "tags": ["SCIE", "新锐2区"],
                    }
                ],
            }

            write_sqlite_index(payload, sqlite_path)

            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_DB": sqlite_path}, clear=False):
                result = metrics._get_journal_index_info("Environmental Pollution", "0269-7491")

            self.assertEqual(result["impact_factor"], "8.8")
            self.assertEqual(result["jcr_quartile"], "Q1")
            self.assertEqual(result["cas_partition_2025"], "2区")
            self.assertEqual(result["xinrui_partition_2026"], "2区")
            self.assertEqual(result["sci_type"], "SCIE")

    def test_index_client_uses_bundled_sqlite_without_environment(self):
        with patch.dict(
            os.environ,
            {
                "SCI_SELECT_JOURNAL_INDEX_DB": "",
                "SCI_SELECT_JOURNAL_INDEX_PATH": "",
                "SCI_SELECT_JOURNAL_INDEX_URL": "",
            },
            clear=False,
        ):
            result = metrics._get_journal_index_info("Environmental Pollution", "0269-7491")

        self.assertIsNotNone(result)
        self.assertEqual(result["cas_partition_2025"], "2区")
        self.assertEqual(result["xinrui_partition_2026"], "2区")
        self.assertTrue(result["nature_index"])
        self.assertEqual(result["nature_index_year"], 2026)

    def test_bundled_index_treats_ampersand_and_and_as_the_same_title(self):
        with patch.dict(
            os.environ,
            {
                "SCI_SELECT_JOURNAL_INDEX_DB": "",
                "SCI_SELECT_JOURNAL_INDEX_PATH": "",
                "SCI_SELECT_JOURNAL_INDEX_URL": "",
            },
            clear=False,
        ):
            result = metrics._get_journal_index_info(
                "Renewable and Sustainable Energy Reviews",
                "1364-0321",
            )

        self.assertIsNotNone(result)
        self.assertEqual(result["cas_partition_2025"], "1区")
        self.assertEqual(result["xinrui_partition_2026"], "1区")
        self.assertNotIn("impact_factor", result)

    def test_build_journal_index_accepts_generic_jcr_2025_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            jcr_path = os.path.join(tmpdir, "jcr_2025.xlsx")

            jcr_book = Workbook()
            jcr_ws = jcr_book.active
            jcr_ws.append(["Journal name", "ISSN", "eISSN", "JIF", "JIF Quartile", "JCR Category", "Edition"])
            jcr_ws.append(["Environmental Pollution", "0269-7491", "1873-6424", "8.8", "Q1", "ENVIRONMENTAL SCIENCES", "SCIE"])
            jcr_ws.append(["Environmental Pollution", "0269-7491", "1873-6424", "8.8", "Q2", "PUBLIC HEALTH", "SCIE"])
            jcr_book.save(jcr_path)

            payload = build_index(jcr_file=jcr_path)
            row = payload["journals"][0]

            self.assertEqual(row["title"], "Environmental Pollution")
            self.assertEqual(row["jif_2025"], "8.8")
            self.assertEqual(row["jcr_quartile_2025"], "Q1")
            self.assertEqual(row["jcr_release_year"], 2026)
            self.assertEqual(row["jcr_data_year"], 2025)
            self.assertIn("SCIE", row["tags"])
            self.assertEqual(len(row["jcr_categories"]), 2)

    def test_build_index_never_cross_merges_journals_on_a_bad_issn(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            cas_path = os.path.join(tmpdir, "cas.xlsx")
            xinrui_path = os.path.join(tmpdir, "xinrui.xlsx")
            jcr_path = os.path.join(tmpdir, "jcr.xlsx")

            cas_book = Workbook()
            cas_ws = cas_book.active
            cas_ws.append(["Journal", "分区"])
            cas_ws.append(["Advanced Materials", "1区"])
            cas_ws.append(["Food Chemistry", "1区"])
            cas_book.save(cas_path)

            xinrui_book = Workbook()
            xinrui_ws = xinrui_book.active
            xinrui_ws.append(["期刊名称", "新锐分区", "issn1"])
            xinrui_ws.append(["Advanced Materials", "1区", "0308-8146"])
            xinrui_ws.append(["Food Chemistry", "1区", "0268-0939"])
            xinrui_book.save(xinrui_path)

            jcr_book = Workbook()
            jcr_ws = jcr_book.active
            jcr_ws.append(["Journal name", "ISSN", "JIF", "JIF Quartile", "Edition"])
            jcr_ws.append(["Advanced Materials", "0935-9648", "26.8", "Q1", "SCIE"])
            jcr_ws.append(["Food Chemistry", "0308-8146", "9.8", "Q1", "SCIE"])
            jcr_book.save(jcr_path)

            payload = build_index(
                cas_2025_xlsx=cas_path,
                xinrui_2026_xlsx=xinrui_path,
                jcr_file=jcr_path,
            )
            rows = {row["title"].lower(): row for row in payload["journals"]}

            self.assertEqual(len(rows), 2)
            self.assertEqual(rows["advanced materials"]["issn"], "0935-9648")
            self.assertEqual(rows["advanced materials"]["jif_2025"], "26.8")
            self.assertEqual(rows["food chemistry"]["issn"], "0308-8146")
            self.assertEqual(rows["food chemistry"]["jif_2025"], "9.8")

    def test_build_journal_index_accepts_showjcr_jcr_2025_csv(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            jcr_path = os.path.join(tmpdir, "JCR2025-UTF8.csv")
            with open(jcr_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(
                    [
                        "Journal",
                        "ISSN",
                        "EISSN",
                        "Web of Science",
                        "IF(2025)",
                        "Category_1",
                        "IF Quartile(2025)_1",
                        "IF Rank(2025)_1",
                        "Category_2",
                        "IF Quartile(2025)_2",
                        "IF Rank(2025)_2",
                    ]
                )
                writer.writerow(
                    [
                        "ENVIRONMENTAL POLLUTION",
                        "0269-7491",
                        "1873-6424",
                        "SCIE",
                        "7.2",
                        "ENVIRONMENTAL SCIENCES",
                        "Q1",
                        "45/378",
                        "PUBLIC HEALTH, ENVIRONMENTAL & OCCUPATIONAL",
                        "Q2",
                        "78/408",
                    ]
                )

            payload = build_index(jcr_file=jcr_path)
            row = payload["journals"][0]

            self.assertEqual(row["jif_2025"], "7.2")
            self.assertEqual(row["jcr_quartile_2025"], "Q1")
            self.assertIn("SCIE", row["tags"])
            self.assertEqual(len(row["jcr_categories"]), 2)
            self.assertEqual(row["jcr_categories"][0]["rank"], "45/378")

    def test_build_journal_index_accepts_nature_index_faq_html(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            ni_path = os.path.join(tmpdir, "nature-index-faq.html")
            with open(ni_path, "w", encoding="utf-8") as f:
                f.write(
                    """
                    <h2 id="journals">12. What publications are in the Nature Index?</h2>
                    <ul>
                      <li><i>Environmental Pollution</i> <span>(1557 articles)</span></li>
                    </ul>
                    <ul>
                      <li><i>IEEE/RSJ International Conference on Intelligent Robots and Systems (IROS)</i> <span>(1958 articles)</span></li>
                    </ul>
                    """
                )

            payload = build_index(nature_index_file=ni_path)
            rows = {row["title"]: row for row in payload["journals"]}

            self.assertTrue(rows["Environmental Pollution"]["nature_index"])
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows["Environmental Pollution"]["nature_index_year"], 2026)
            self.assertEqual(rows["Environmental Pollution"]["nature_index_articles"], 1557)
            self.assertEqual(rows["Environmental Pollution"]["nature_index_publication_type"], "journal")
            self.assertEqual(
                rows["IEEE/RSJ International Conference on Intelligent Robots and Systems (IROS)"]["nature_index_publication_type"],
                "conference_proceeding",
            )

    def test_index_client_reads_current_jcr_2025_fields(self):
        payload = {
            "journals": [
                {
                    "title": "ENVIRONMENTAL POLLUTION",
                    "issn": "0269-7491",
                    "eissn": "1873-6424",
                    "jif_2025": "8.8",
                    "jcr_release_year": 2026,
                    "jcr_data_year": 2025,
                    "jcr_quartile_2025": "Q1",
                    "cas_2025": "2区",
                    "xuankan_2026": "2区",
                    "tags": ["SCIE"],
                }
            ]
        }
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as f:
            json.dump(payload, f, ensure_ascii=False)
            index_path = f.name

        try:
            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_PATH": index_path}):
                result = metrics._get_journal_index_info("Environmental Pollution", "0269-7491")
        finally:
            os.unlink(index_path)

        self.assertEqual(result["impact_factor"], "8.8")
        self.assertEqual(result["if_year"], "2025")
        self.assertEqual(result["jcr_release_year"], 2026)
        self.assertEqual(result["jcr_data_year"], 2025)
        self.assertEqual(result["jcr_quartile"], "Q1")
        self.assertEqual(result["xinrui_partition_2026"], "2区")

    def test_index_client_rejects_issn_hit_for_a_different_journal(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            sqlite_path = os.path.join(tmpdir, "journal-index.sqlite")
            payload = {
                "meta": {"schema": "sci-select-journal-index-v1"},
                "journals": [
                    {
                        "title": "Advanced Materials",
                        "issn": "0308-8146",
                        "cas_2025": "1区",
                    },
                    {
                        "title": "Food Chemistry",
                        "issn": "9999-9999",
                        "cas_2025": "2区",
                    },
                ],
            }
            write_sqlite_index(payload, sqlite_path)

            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_DB": sqlite_path}, clear=False):
                result = metrics._get_journal_index_info("Food Chemistry", "0308-8146")

            self.assertEqual(result["name"], "Food Chemistry")
            self.assertEqual(result["cas_partition_2025"], "2区")

    def test_index_client_reads_nature_index_fields(self):
        payload = {
            "journals": [
                {
                    "title": "ENVIRONMENTAL POLLUTION",
                    "issn": "0269-7491",
                    "nature_index": True,
                    "nature_index_year": 2026,
                    "nature_index_articles": 1557,
                    "nature_index_publication_type": "journal",
                    "tags": ["Nature Index"],
                }
            ]
        }
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as f:
            json.dump(payload, f, ensure_ascii=False)
            index_path = f.name

        try:
            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_PATH": index_path}):
                result = metrics._get_journal_index_info("Environmental Pollution", "0269-7491")
        finally:
            os.unlink(index_path)

        self.assertTrue(result["nature_index"])
        self.assertEqual(result["nature_index_year"], 2026)
        self.assertEqual(result["nature_index_articles"], 1557)
        self.assertIn("NI=2026", metrics.format_metrics_line(result))

    def test_letpub_search_keeps_rating_separate_from_real_time_if(self):
        html = """
        <table class="table_yjfx">
          <tr><td>ISSN</td><td>期刊名</td><td>综合评分</td><td>期刊指标</td><td>新锐</td><td>领域</td><td>SCI</td></tr>
          <tr>
            <td>0935-9648</td>
            <td><a href="?journalid=209">ADVANCED MATERIALS</a></td>
            <td>9.5</td>
            <td>实时IF: 29.1 h-index: 447 CiteScore: 42.0</td>
            <td>1区</td><td>材料科学</td><td>SCI SCIE</td>
          </tr>
        </table>
        """

        result = letpub.parse_search_results(html)["journals"][0]

        self.assertEqual(result["rating"], "9.5")
        self.assertEqual(result["real_time_if"], "29.1")
        self.assertEqual(result["xinrui_partition_2026"], "1区")
        self.assertNotIn("partition", result)
        self.assertNotIn("impact_factor", result)

    def test_legacy_recommend_never_treats_ambiguous_partition_as_xinrui(self):
        with patch.object(
            legacy_recommend,
            "search_candidates",
            return_value=[
                {
                    "name": "Legacy Journal",
                    "partition": "1区",
                    "field": "environmental science",
                    "sci_type": "SCIE",
                }
            ],
        ):
            results = legacy_recommend.recommend(
                text="environmental science",
                enrich_details=False,
                max_candidates=1,
            )

        self.assertFalse(results[0].get("xinrui_partition_2026"))

    def test_letpub_lookup_prefers_exact_issn_and_validates_identity(self):
        search_result = {
            "journals": [
                {
                    "name": "Advanced Materials Interfaces",
                    "issn": "2196-7350",
                    "journal_id": "wrong",
                },
                {
                    "name": "ADVANCED MATERIALS",
                    "issn": "0935-9648",
                    "journal_id": "209",
                    "rating": "9.5",
                    "real_time_if": "29.1",
                },
            ]
        }
        detail = {
            "name": "ADVANCED MATERIALS",
            "issn": "0935-9648",
            "field": "Materials Science",
        }

        with patch.object(letpub, "advanced_search", return_value=search_result), \
            patch.object(letpub, "autocomplete_journal", return_value=[]), \
            patch.object(letpub, "get_journal_detail", return_value=detail) as get_detail:
            result = letpub.lookup_journal("Advanced Materials", "0935-9648")

        get_detail.assert_called_once_with("209")
        self.assertEqual(result["issn"], "0935-9648")
        self.assertEqual(result["rating"], "9.5")
        self.assertNotIn("impact_factor", result)

    def test_letpub_detail_parses_last_number_from_dated_real_time_if(self):
        html = """
        <table>
          <tr><td>实时影响因子</td><td>截止2026年5月06日：27.17</td></tr>
          <tr><td>五年IF （数据来源于网友提供）</td><td>29.596 数据由网友提供</td></tr>
        </table>
        """

        detail = letpub.parse_detail_page(html)

        self.assertEqual(detail["real_time_if"], "27.17")
        self.assertEqual(detail["five_year_if"], "29.596")

    def test_infers_english_groundwater_isotope_categories(self):
        profile = infer_paper_profile(
            "Groundwater nitrate source identification using stable isotopes "
            "and hydrochemistry in an agricultural watershed."
        )

        categories = {(c["category1"], c["category2"]) for c in profile["categories"]}

        self.assertIn(("环境科学与生态学", "水资源"), categories)
        self.assertIn(("地球科学", "地球化学与地球物理"), categories)
        self.assertNotEqual(
            profile["categories"],
            [{"category1": "环境科学与生态学", "category2": "环境科学"}],
        )

    def test_infers_medical_ai_categories(self):
        profile = infer_paper_profile(
            "A deep learning radiomics model predicts lung cancer prognosis "
            "from CT imaging and clinical records."
        )

        categories = {(c["category1"], c["category2"]) for c in profile["categories"]}

        self.assertIn(("医学", "肿瘤学"), categories)
        self.assertIn(("计算机科学", "计算机：人工智能"), categories)

    def test_applied_method_terms_do_not_dominate_domain_terms(self):
        profile = infer_paper_profile(
            "Machine learning model for crop irrigation scheduling in agricultural farmland."
        )

        categories = [(c["category1"], c["category2"]) for c in profile["categories"]]

        self.assertEqual(categories[0], ("农林科学", "农业综合"))
        self.assertNotEqual(categories[0], ("计算机科学", "计算机：人工智能"))
        self.assertIn("machine learning", profile["methods"])

    def test_ranking_uses_fit_quality_and_risk_flags(self):
        profile = infer_paper_profile(
            "Groundwater nitrate source identification using stable isotopes "
            "and hydrochemistry in an agricultural watershed."
        )
        records = [
            {
                "name": "Journal of Hydrology",
                "impact_factor": "6.3",
                "partition": "1区",
                "sci_type": "SCIE",
                "h_index": 198,
                "field": "水资源; 地球化学与地球物理",
                "speed": "网友分享经验：平均8.3个月",
                "_sources": ["letpub", "openalex"],
            },
            {
                "name": "Broad Environmental Letters",
                "impact_factor": "8.0",
                "partition": "1区",
                "sci_type": "SCIE",
                "h_index": 80,
                "field": "环境科学",
                "_sources": ["letpub", "openalex"],
            },
            {
                "name": "Emerging Water Reports",
                "impact_factor": "2.1",
                "partition": "4区",
                "sci_type": "ESCI",
                "field": "水资源",
                "_sources": ["letpub"],
                "_source_errors": {"openalex": "timeout"},
            },
        ]

        ranked = rank_metric_records(profile, records)

        self.assertEqual(ranked[0]["name"], "Journal of Hydrology")
        self.assertEqual(ranked[0]["tier"], "推荐")
        self.assertEqual(ranked[-1]["tier"], "谨慎")
        self.assertIn("OpenAlex未获取", ranked[-1]["data_notes"])

    def test_report_does_not_expose_login_or_comment_flow(self):
        profile = infer_paper_profile("groundwater isotope hydrochemistry")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Journal of Hydrology",
                    "impact_factor": "6.3",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "水资源; 地球化学与地球物理",
                    "_sources": ["letpub"],
                    "_source_errors": {"openalex": "SSL error"},
                }
            ],
        )

        report = format_selection_report(profile, ranked)

        self.assertIn("sci-select", report)
        self.assertIn("OpenAlex未获取", report)
        self.assertNotIn("评论", report)
        self.assertNotIn("登录", report)
        self.assertNotIn("cookie", report.lower())
        self.assertNotIn("comments_mode", inspect.signature(selector.select_journals).parameters)

    def test_openalex_source_matching_rejects_unrelated_name(self):
        source = {
            "display_name": "Journal of Hydrology",
            "issn_l": "0022-1694",
            "issn": ["0022-1694"],
        }

        self.assertTrue(metrics._openalex_source_matches(source, "Journal of Hydrology"))
        self.assertTrue(metrics._openalex_source_matches(source, "J Hydrol", "00221694"))
        self.assertFalse(metrics._openalex_source_matches(source, "Nature"))

    def test_openalex_source_matching_handles_missing_lists(self):
        source = {
            "display_name": "Journal of Hydrology",
            "issn_l": None,
            "issn": None,
            "alternate_titles": None,
        }

        self.assertTrue(metrics._openalex_source_matches(source, "Journal of Hydrology"))

    def test_letpub_detail_extracts_public_xinrui_partition(self):
        html = """
        <table>
          <tr>
            <td>《新锐期刊分区表》 （ 2026年3月发布 ）</td>
            <td>趋势图</td>
            <td>地球科学 3区 1区 4区</td>
            <td>WATER RESOURCES 水资源 2区 3区 1区</td>
            <td>WATER RESOURCES 水资源</td>
            <td>2区 3区 1区</td>
            <td>是</td>
            <td>N/A</td>
            <td>期刊分区表 （ 2025年3月升级版 ）</td>
            <td>地球科学 3区 1区 4区</td>
            <td>WATER RESOURCES 水资源 4区 2区 1区</td>
            <td>是</td>
            <td>否</td>
          </tr>
        </table>
        """

        detail = letpub.parse_detail_page(html)

        self.assertEqual(detail["xinrui_partition_2026"], "3区")
        self.assertEqual(detail["xinrui_2026"]["分区"], "3区")
        self.assertTrue(detail["xinrui_2026"]["Top期刊"])
        self.assertFalse(detail["xinrui_2026"]["综述期刊"])

    def test_letpub_xinrui_partition_ignores_hidden_decoy_spans(self):
        html = """
        <table>
          <tr>
            <td>《新锐期刊分区表》 （ 2026年3月发布 ）</td>
            <td colspan="2">
              <table>
                <tr>
                  <th>大类学科</th><th>小类学科</th><th>Top期刊</th><th>综述期刊</th>
                </tr>
                <tr>
                  <td>
                    环境科学与生态学
                    <span style="display:none">4区</span>
                    <span>2区</span>
                    <span style="display:none">3区</span>
                  </td>
                  <td>
                    <table>
                      <tr>
                        <td>ENVIRONMENTAL SCIENCES<br/>环境科学</td>
                        <td>
                          <span style="display:none">1区</span>
                          <span style="display:none">4区</span>
                          <span>2区</span>
                        </td>
                      </tr>
                    </table>
                  </td>
                  <td>是</td>
                  <td>N/A</td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
        """

        detail = letpub.parse_detail_page(html)

        self.assertEqual(detail["xinrui_partition_2026"], "2区")
        self.assertEqual(detail["xinrui_2026"]["小类学科"], "ENVIRONMENTAL SCIENCES 环境科学 2区")
        self.assertTrue(detail["xinrui_2026"]["Top期刊"])
        self.assertFalse(detail["xinrui_2026"]["综述期刊"])

    def test_metrics_prefers_letpub_xinrui_before_api_lookup(self):
        letpub_detail = {
            "issn": "0022-1694",
            "impact_factor": "6.3",
            "ch_sci_2025": {"分区": "1区"},
            "xinrui_partition_2026": "3区",
        }

        result = metrics._merge_letpub_metrics({"name": "Journal of Hydrology", "_sources": []}, letpub_detail)

        self.assertEqual(result["cas_partition_2025"], "1区")
        self.assertEqual(result["xinrui_partition_2026"], "3区")

    def test_metrics_ignores_stale_cache_schema(self):
        stale_cache = {
            "Environmental Pollution": {
                "name": "Environmental Pollution",
                "_sources": ["letpub", "openalex"],
                "_source_errors": {},
                "_cached_at": 9999999999,
                "_cache_schema_version": metrics.CACHE_SCHEMA_VERSION - 1,
                "xinrui_partition_2026": "1区",
            }
        }
        letpub_detail = {
            "issn": "0269-7491",
            "impact_factor": "7.2",
            "ch_sci_2025": {"分区": "2区"},
            "xinrui_partition_2026": "2区",
            "sci_type": "SCI SCIE",
        }

        with patch.object(metrics, "_load_cache", return_value=stale_cache), \
            patch.object(metrics, "_save_cache"), \
            patch.object(metrics, "_get_letpub_info", return_value=letpub_detail), \
            patch.object(metrics, "_get_openalex_info", return_value=None):
            result = metrics.get_journal_metrics("Environmental Pollution")

        self.assertEqual(result["xinrui_partition_2026"], "2区")

    def test_selection_metrics_skip_letpub_when_local_index_has_core_fields(self):
        local_record = {
            "name": "Natural Hazards",
            "issn": "0921-030X",
            "impact_factor": "3.7",
            "jcr_quartile": "Q2",
            "cas_partition_2025": "3区",
            "xinrui_partition_2026": "3区",
            "sci_type": "SCIE",
        }
        with patch.object(metrics, "_load_cache", return_value={}), \
            patch.object(metrics, "_save_cache"), \
            patch.object(metrics, "_get_journal_index_info", return_value=local_record), \
            patch.object(metrics, "_get_letpub_info") as get_letpub, \
            patch.object(metrics, "_get_openalex_info", return_value=None), \
            patch.object(metrics, "_get_xinrui_info", return_value=None):
            result = metrics.get_journal_metrics(
                "Natural Hazards",
                use_cache=False,
                source_mode="selection",
            )

        get_letpub.assert_not_called()
        self.assertIn("letpub", result["_skipped_sources"])
        self.assertEqual(result["cas_partition_2025"], "3区")

    def test_selection_uses_letpub_categories_only_when_literature_recall_is_short(self):
        literature_candidates = [
            {
                "name": f"Evidence Journal {index}",
                "similar_works_count": 3,
                "query_coverage": 2,
                "literature_evidence_score": 20,
                "publication_precedents": [{"title": f"Related {index}", "year": 2025}],
                "_sources": ["openalex-works"],
            }
            for index in range(4)
        ]
        metric_record = {
            "impact_factor": "4.0",
            "jcr_quartile": "Q2",
            "sci_type": "SCIE",
            "_sources": ["journal-index"],
        }
        with patch.object(
            selector,
            "discover_journal_candidates",
            return_value={"queries": ["query"], "candidates": literature_candidates, "errors": []},
        ), patch.object(selector, "search_candidates") as category_search, patch.object(
            selector,
            "get_journal_metrics",
            side_effect=lambda name, **kwargs: {"name": name, **metric_record},
        ), patch.object(selector.time, "sleep"):
            bundle = selector.select_journals(
                "cross-disciplinary manuscript abstract",
                paper_profile={
                    "direction_summary": "structured direction",
                    "research_object": "research object",
                    "search_queries": ["query"],
                },
                max_candidates=3,
            )

        category_search.assert_not_called()
        self.assertEqual(len(bundle["results"]), 3)

    def test_metrics_ignores_incomplete_current_cache_record(self):
        incomplete_cache = {
            "Environmental Pollution": {
                "name": "Environmental Pollution",
                "_sources": ["letpub", "openalex"],
                "_source_errors": {},
                "_cached_at": 9999999999,
                "_cache_schema_version": metrics.CACHE_SCHEMA_VERSION,
                "issn": "",
                "impact_factor": None,
                "partition": "",
                "cas_partition_2025": None,
                "xinrui_partition_2026": None,
                "sci_type": "",
            }
        }
        letpub_detail = {
            "issn": "0269-7491",
            "impact_factor": "8.8",
            "ch_sci_2025": {"分区": "2区"},
            "xinrui_partition_2026": "2区",
            "_sci_type": "SCI SCIE",
        }
        openalex_detail = {
            "h_index": 194,
            "i10_index": 5000,
            "2yr_mean_citedness": 15.1,
            "cited_by_count": 1246754,
            "works_count": 42000,
            "oa_works_count": 12000,
            "is_oa": False,
            "is_in_doaj": False,
            "apc_usd": None,
            "host_organization": "Elsevier",
        }

        with patch.object(metrics, "_load_cache", return_value=incomplete_cache), \
            patch.object(metrics, "_save_cache"), \
            patch.object(metrics, "_get_journal_index_info", return_value=None), \
            patch.object(metrics, "_get_letpub_info", return_value=letpub_detail), \
            patch.object(metrics, "_get_openalex_info", return_value=openalex_detail), \
            patch.object(metrics.time, "sleep"):
            result = metrics.get_journal_metrics("Environmental Pollution")

        self.assertEqual(result["issn"], "0269-7491")
        self.assertEqual(result["impact_factor"], "8.8")
        self.assertEqual(result["cas_partition_2025"], "2区")
        self.assertEqual(result["xinrui_partition_2026"], "2区")
        self.assertIn("SCIE", result["sci_type"])

    def test_metrics_can_use_local_journal_index_for_stable_partitions(self):
        payload = {
            "meta": {"generated_at": "2026-06-17T14:26:31", "source": "test-index"},
            "journals": [
                {
                    "title": "ENVIRONMENTAL POLLUTION",
                    "issn": "0269-7491",
                    "eissn": "1873-6424",
                    "if_2023": 7.2,
                    "if_year": "2025",
                    "jcr_quartile": "Q1",
                    "cas_2025": "2区",
                    "xuankan_2026": "2区",
                    "tags": ["SCIE", "Q1", "新锐2区"],
                }
            ],
        }
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as f:
            json.dump(payload, f, ensure_ascii=False)
            index_path = f.name

        try:
            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_PATH": index_path}), \
                patch.object(metrics, "_load_cache", return_value={}), \
                patch.object(metrics, "_save_cache"), \
                patch.object(metrics, "_get_letpub_info", return_value=None), \
                patch.object(metrics, "_get_openalex_info", return_value=None):
                result = metrics.get_journal_metrics("Environmental Pollution", use_cache=False)
        finally:
            os.unlink(index_path)

        self.assertEqual(result["issn"], "0269-7491")
        self.assertEqual(result["impact_factor"], 7.2)
        self.assertEqual(result["cas_partition_2025"], "2区")
        self.assertEqual(result["xinrui_partition_2026"], "2区")
        self.assertIn("journal-index", result["_sources"])
        self.assertIn("2026新锐=2区", metrics.format_metrics_line(result))

    def test_metrics_keeps_journal_index_partition_when_letpub_conflicts(self):
        payload = {
            "journals": [
                {
                    "title": "JOURNAL OF HAZARDOUS MATERIALS",
                    "issn": "0304-3894",
                    "cas_2025": "1区",
                    "xuankan_2026": "1区",
                    "tags": ["SCIE"],
                }
            ]
        }
        letpub_detail = {
            "issn": "0304-3894",
            "impact_factor": "10.6",
            "ch_sci_2025": {"分区": "2区"},
            "xinrui_partition_2026": "3区",
            "_sci_type": "SCIE",
        }
        with tempfile.NamedTemporaryFile("w", encoding="utf-8", suffix=".json", delete=False) as f:
            json.dump(payload, f, ensure_ascii=False)
            index_path = f.name

        try:
            with patch.dict(os.environ, {"SCI_SELECT_JOURNAL_INDEX_PATH": index_path}), \
                patch.object(metrics, "_load_cache", return_value={}), \
                patch.object(metrics, "_save_cache"), \
                patch.object(metrics, "_get_letpub_info", return_value=letpub_detail), \
                patch.object(metrics, "_get_openalex_info", return_value=None):
                result = metrics.get_journal_metrics("Journal of Hazardous Materials", use_cache=False)
        finally:
            os.unlink(index_path)

        self.assertEqual(result["cas_partition_2025"], "1区")
        self.assertEqual(result["xinrui_partition_2026"], "1区")
        self.assertIn("分区来源冲突需复核", "；".join(result["status_notes"]))

    def test_known_removed_wos_journal_overrides_stale_scie_status(self):
        record = {
            "name": "The Science of the Total Environment",
            "sci_type": "SCIE",
            "warning": False,
            "_sources": ["letpub"],
        }

        metrics._apply_known_status_overrides(record)

        self.assertEqual(record["wos_status"], "removed")
        self.assertEqual(record["sci_type"], "WOS_REMOVED")
        self.assertTrue(record["warning"])
        self.assertIn("WoS已移除", metrics.format_metrics_line(record))

    def test_removed_wos_status_is_not_recommended(self):
        profile = infer_paper_profile("environmental pollution water quality")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Science of the Total Environment",
                    "impact_factor": "8.0",
                    "partition": "1区",
                    "sci_type": "WOS_REMOVED",
                    "wos_status": "removed",
                    "field": "环境科学; 水资源",
                    "_sources": ["letpub"],
                }
            ],
        )

        self.assertEqual(ranked[0]["tier"], "不推荐")
        self.assertIn("Web of Science", "；".join(ranked[0]["risk_reasons"]))

    def test_metrics_line_labels_2025_cas_and_2026_xinrui_partitions(self):
        line = metrics.format_metrics_line(
            {
                "name": "Journal of Hydrology",
                "sci_type": "SCIE",
                "impact_factor": "6.3",
                "cas_partition_2025": "1区",
                "xinrui_partition_2026": "2区Top",
            }
        )

        self.assertIn("2025中科院=1区", line)
        self.assertIn("2026新锐=2区Top", line)
        self.assertNotIn("分区=1区", line)

    def test_metrics_line_never_treats_missing_coverage_as_non_sci(self):
        line = metrics.format_metrics_line(
            {
                "name": "Advanced Materials",
                "real_time_if": "27.17",
                "cas_partition_2025": "1区",
                "xinrui_partition_2026": "1区",
                "_sources": ["journal-index", "letpub"],
            }
        )

        self.assertIn("收录未获取", line)
        self.assertIn("实时IF≈27.17(非JIF)", line)
        self.assertNotIn("非SCI", line)

    def test_matrix_report_shows_cas_2025_and_xinrui_2026_columns(self):
        profile = infer_paper_profile("groundwater isotope hydrochemistry")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Journal of Hydrology",
                    "impact_factor": "6.3",
                    "cas_partition_2025": "1区",
                    "xinrui_partition_2026": "2区Top",
                    "sci_type": "SCIE",
                    "field": "水资源; 地球化学与地球物理",
                    "_sources": ["letpub", "openalex", "xinrui"],
                }
            ],
        )

        matrix = format_selection_matrix(profile, ranked)

        self.assertIn("| 期刊 | 候选状态 | 方向证据 | 期刊层级 | IF | 2025中科院 | 2026新锐 | 收录 |", matrix)
        self.assertIn("| Journal of Hydrology | 优先核验 |", matrix)
        self.assertIn("| 1区 | 2区Top | SCIE |", matrix)

    def test_letpub_xinrui_partition_is_not_reported_missing(self):
        profile = infer_paper_profile("groundwater isotope hydrochemistry")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Journal of Hydrology",
                    "impact_factor": "6.3",
                    "cas_partition_2025": "1区",
                    "xinrui_partition_2026": "3区",
                    "sci_type": "SCIE",
                    "field": "水资源; 地球化学与地球物理",
                    "_sources": ["letpub", "openalex"],
                }
            ],
        )

        self.assertNotIn("2026新锐分区未获取", ranked[0]["data_notes"])
        self.assertIn("LetPub新锐", ranked[0]["data_status"])

    def test_matrix_report_shows_decision_table(self):
        profile = infer_paper_profile("groundwater isotope hydrochemistry")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Journal of Hydrology",
                    "impact_factor": "6.3",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "水资源; 地球化学与地球物理",
                    "speed": "网友分享经验：平均8.3个月",
                    "_sources": ["letpub", "openalex"],
                    "h_index": 198,
                    "is_oa": False,
                }
            ],
        )

        matrix = format_selection_matrix(profile, ranked)

        self.assertIn("| 期刊 | 候选状态 | 方向证据 | 期刊层级 | IF |", matrix)
        self.assertIn("Journal of Hydrology", matrix)
        self.assertIn("SCIE", matrix)
        self.assertIn("平均8.3个月", matrix)

    def test_select_journals_can_strictly_filter_xinrui_partition(self):
        candidates = [
            {
                "name": "Journal of Cleaner Production",
                "impact_factor": "9.8",
                "partition": "1区",
                "sci_type": "SCIE",
                "field": "环境科学",
            },
            {
                "name": "Environmental Pollution",
                "impact_factor": "8.8",
                "partition": "2区",
                "sci_type": "SCIE",
                "field": "环境科学",
            },
        ]
        metrics_by_name = {
            "Journal of Cleaner Production": {
                "name": "Journal of Cleaner Production",
                "impact_factor": "9.8",
                "cas_partition_2025": "1区",
                "xinrui_partition_2026": "1区",
                "sci_type": "SCIE",
                "field": "环境科学",
                "_sources": ["letpub", "openalex"],
            },
            "Environmental Pollution": {
                "name": "Environmental Pollution",
                "impact_factor": "8.8",
                "cas_partition_2025": "2区",
                "xinrui_partition_2026": "2区",
                "sci_type": "SCIE",
                "field": "环境科学",
                "_sources": ["letpub", "openalex"],
            },
        }

        with patch.object(selector, "search_candidates", return_value=candidates), \
            patch.object(selector, "discover_specialist_candidates", return_value=[]), \
            patch.object(selector, "get_journal_metrics", side_effect=lambda name, **kwargs: metrics_by_name[name]), \
            patch.object(selector.time, "sleep"):
            bundle = selector.select_journals(
                "groundwater nitrate environmental pollution",
                categories=[{"category1": "环境科学与生态学", "category2": "环境科学", "score": 1}],
                xinrui_partition="1区",
                max_candidates=5,
                use_literature_evidence=False,
            )

        self.assertEqual(
            [item["name"] for item in bundle["results"]],
            ["Journal of Cleaner Production"],
        )
        self.assertEqual(bundle["results"][0]["xinrui_partition_2026"], "1区")

    def test_xinrui_filter_keeps_literature_candidate_until_metrics_are_fetched(self):
        literature_candidate = {
            "name": "Evidence Journal",
            "similar_works_count": 3,
            "query_coverage": 2,
            "literature_evidence_score": 18,
            "publication_precedents": [{"title": "Related work", "year": 2025}],
            "_sources": ["openalex-works"],
        }
        fetched = {
            "name": "Evidence Journal",
            "xinrui_partition_2026": "1区",
            "sci_type": "SCIE",
            "_sources": ["journal-index"],
        }
        with patch.object(
            selector,
            "discover_journal_candidates",
            return_value={"queries": ["query"], "candidates": [literature_candidate], "errors": []},
        ), patch.object(selector, "discover_specialist_candidates", return_value=[]), \
            patch.object(selector, "search_candidates", return_value=[]), \
            patch.object(selector, "get_journal_metrics", return_value=fetched):
            bundle = selector.select_journals(
                "paper abstract",
                paper_profile={"search_queries": ["query"]},
                xinrui_partition="1区",
                max_candidates=3,
                request_delay=0,
            )

        self.assertEqual([row["name"] for row in bundle["results"]], ["Evidence Journal"])
        self.assertEqual(bundle["results"][0]["similar_works_count"], 3)

    def test_candidate_label_uses_current_xinrui_journal_level(self):
        profile = infer_paper_profile("groundwater nitrate environmental science")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Current XinRui One",
                    "impact_factor": "6.0",
                    "cas_partition_2025": "2区",
                    "xinrui_partition_2026": "1区",
                    "sci_type": "SCIE",
                    "field": "环境科学; 水资源",
                    "_sources": ["letpub", "openalex"],
                }
            ],
        )

        banded = assign_candidate_labels(ranked)

        self.assertEqual(banded[0]["candidate_label"], "高位候选")
        self.assertEqual(banded[0]["submission_band"], "高位候选")
        self.assertIn("2026新锐1区", "；".join(banded[0]["quality_reasons"]))

    def test_candidate_groups_are_interleaved_across_categories(self):
        groups = [
            [{"name": "Water A"}, {"name": "Water B"}, {"name": "Water C"}],
            [{"name": "Geochem A"}, {"name": "Geochem B"}],
            [{"name": "Geology A"}],
        ]

        candidates = interleave_candidate_groups(groups, limit=5)

        self.assertEqual(
            [c["name"] for c in candidates],
            ["Water A", "Geochem A", "Geology A", "Water B", "Geochem B"],
        )

    def test_review_journal_is_cautious_for_non_review_paper(self):
        profile = infer_paper_profile("groundwater nitrate hydrochemistry field sampling")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "REVIEWS OF GEOPHYSICS",
                    "impact_factor": "37.3",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "地球化学与地球物理",
                    "_sources": ["letpub", "openalex"],
                }
            ],
        )

        self.assertEqual(ranked[0]["tier"], "谨慎")
        self.assertIn("综述型期刊", "；".join(ranked[0]["risk_reasons"]))

    def test_legacy_submission_bands_expose_journal_levels_only(self):
        profile = infer_paper_profile("groundwater nitrate hydrochemistry")
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Ambition Journal",
                    "impact_factor": "12",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "水资源",
                    "_sources": ["letpub"],
                },
                {
                    "name": "Solid Journal",
                    "impact_factor": "5.5",
                    "partition": "2区",
                    "sci_type": "SCIE",
                    "field": "水资源",
                    "_sources": ["letpub"],
                },
                {
                    "name": "Safe Journal",
                    "impact_factor": "3.2",
                    "partition": "3区",
                    "sci_type": "SCIE",
                    "field": "水资源",
                    "_sources": ["letpub"],
                },
            ],
        )

        banded = assign_submission_bands(ranked)
        report = format_selection_report(profile, banded)

        self.assertEqual(
            {item["name"]: item["submission_band"] for item in banded},
            {
                "Ambition Journal": "高位候选",
                "Solid Journal": "中位候选",
                "Safe Journal": "常规候选",
            },
        )
        self.assertIn("不评价稿件水平", report)
        self.assertIn("期刊层级", report)
        self.assertNotIn("主投", report)
        self.assertNotIn("稳妥", report)
        self.assertNotIn("保底", report)

    def test_report_warns_when_candidate_recall_is_low_confidence(self):
        profile = {
            "categories": [{"category1": "环境科学与生态学", "category2": "水资源", "score": 3}],
            "matched_terms": ["hydrology"],
            "methods": ["machine learning"],
        }
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "High Impact Broad Methods",
                    "impact_factor": "12",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "Computer Science; Artificial Intelligence",
                    "_sources": ["letpub", "openalex"],
                },
                {
                    "name": "General Environmental Reports",
                    "impact_factor": "6",
                    "partition": "2区",
                    "sci_type": "SCIE",
                    "field": "Environmental Sciences",
                    "_sources": ["letpub"],
                },
            ],
        )

        report = format_selection_report(profile, ranked)

        self.assertIn("候选召回置信度较低", report)
        self.assertIn("不要仅按 IF 或分区决策", report)

    def test_report_exposes_literature_retrieval_degradation(self):
        profile = infer_paper_profile("groundwater nitrate hydrochemistry")
        profile["retrieval_notes"] = ["OpenAlex相似论文检索失败[1]: timeout"]
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Journal of Hydrology",
                    "impact_factor": "7.3",
                    "sci_type": "SCIE",
                    "field": "水资源",
                    "_sources": ["journal-index"],
                }
            ],
        )

        report = format_selection_report(profile, ranked)

        self.assertIn("检索降级", report)
        self.assertIn("OpenAlex相似论文检索失败", report)

    def test_profile_separates_methods_from_topic_evidence(self):
        profile = infer_paper_profile(
            "Machine learning and GIS are used to map flash flood susceptibility "
            "from rainfall, terrain, and land-use data."
        )

        self.assertIn("machine learning", profile["methods"])
        self.assertIn("gis", profile["methods"])
        self.assertIn("flash flood", profile["topic_evidence"])
        self.assertNotEqual(profile["primary_signal"], "machine learning")

    def test_topic_evidence_can_outrank_broad_high_if_match(self):
        profile = infer_paper_profile(
            "Machine learning and GIS are used to map flash flood susceptibility "
            "from rainfall, terrain, and land-use data."
        )
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "High Impact AI Letters",
                    "impact_factor": "15",
                    "partition": "1区",
                    "sci_type": "SCIE",
                    "field": "Computer Science; Artificial Intelligence",
                    "_sources": ["letpub", "openalex"],
                },
                {
                    "name": "Natural Hazards and Earth System Sciences",
                    "impact_factor": "4.6",
                    "partition": "2区",
                    "sci_type": "SCIE",
                    "field": "Natural hazards; flash flood; rainfall; terrain",
                    "_sources": ["letpub", "openalex"],
                },
            ],
        )

        self.assertEqual(ranked[0]["name"], "Natural Hazards and Earth System Sciences")
        self.assertIn("细分主题", "；".join(ranked[0]["fit_reasons"]))

    def test_structured_profile_replaces_keyword_fallback_direction(self):
        fallback = infer_paper_profile(
            "Machine learning for social-media flash-flood susceptibility mapping."
        )
        profile = merge_paper_profile(
            fallback,
            {
                "direction_summary": "山洪易发性与自然灾害风险制图；机器学习仅作为建模方法",
                "primary_field": "natural hazards",
                "specialty": "flash-flood susceptibility mapping",
                "research_object": "flash floods",
                "research_question": "where flash-flood susceptibility is highest",
                "contribution_type": "applied risk-mapping study",
                "target_audience": ["natural-hazard researchers", "risk managers"],
                "methods": ["machine learning", "GIS"],
                "exclusions": ["general artificial intelligence", "social-media studies"],
                "search_queries": [
                    "flash flood susceptibility mapping rainfall terrain",
                    "natural hazard risk mapping machine learning GIS",
                ],
                "positioning": {
                    "level": "moderate",
                    "confidence": "medium",
                    "evidence": ["multi-factor susceptibility comparison"],
                },
            },
        )

        self.assertEqual(profile["profile_source"], "ai-structured")
        self.assertEqual(profile["research_object"], "flash floods")
        self.assertEqual(profile["primary_signal"], "flash floods")
        self.assertEqual(len(profile["search_queries"]), 2)
        self.assertNotIn("general artificial intelligence", profile["target_audience"])

    def test_openalex_similar_works_are_aggregated_as_journal_precedents(self):
        responses = [
            {
                "results": [
                    {
                        "id": "https://openalex.org/W1",
                        "display_name": "Flash-flood susceptibility mapping with terrain controls",
                        "publication_year": 2024,
                        "type": "article",
                        "relevance_score": 100,
                        "primary_location": {
                            "source": {
                                "id": "https://openalex.org/S1",
                                "display_name": "Natural Hazards",
                                "type": "journal",
                                "issn_l": "0921-030X",
                                "issn": ["0921-030X"],
                            }
                        },
                        "primary_topic": {"display_name": "Flood Risk Assessment"},
                    },
                    {
                        "id": "https://openalex.org/W2",
                        "display_name": "Rainfall and terrain controls on flash floods",
                        "publication_year": 2023,
                        "type": "article",
                        "relevance_score": 80,
                        "primary_location": {
                            "source": {
                                "id": "https://openalex.org/S1",
                                "display_name": "Natural Hazards",
                                "type": "journal",
                                "issn_l": "0921-030X",
                                "issn": ["0921-030X"],
                            }
                        },
                        "primary_topic": {"display_name": "Flood Risk Assessment"},
                    },
                ]
            },
            {
                "results": [
                    {
                        "id": "https://openalex.org/W3",
                        "display_name": "Machine learning for natural-hazard risk maps",
                        "publication_year": 2025,
                        "type": "article",
                        "relevance_score": 90,
                        "primary_location": {
                            "source": {
                                "id": "https://openalex.org/S1",
                                "display_name": "Natural Hazards",
                                "type": "journal",
                                "issn_l": "0921-030X",
                                "issn": ["0921-030X"],
                            }
                        },
                        "primary_topic": {"display_name": "Natural Hazard Mapping"},
                    }
                ]
            },
            {
                "results": [
                    {
                        "id": "https://openalex.org/S1",
                        "works_count": 12000,
                        "counts_by_year": [
                            {"year": 2026, "works_count": 800},
                            {"year": 2025, "works_count": 700},
                            {"year": 2024, "works_count": 600},
                            {"year": 2023, "works_count": 500},
                            {"year": 2022, "works_count": 400},
                        ],
                    }
                ]
            },
        ]

        with patch.dict(os.environ, {"OPENALEX_API_KEY": "test-key"}), \
            patch("scripts.similar_works.requests.get") as get:
            mocked_responses = [
                unittest.mock.Mock(status_code=200, json=lambda payload=payload: payload)
                for payload in responses
            ]
            for response in mocked_responses:
                response.raise_for_status.return_value = None
            get.side_effect = mocked_responses
            result = discover_journal_candidates(
                [
                    "flash flood susceptibility mapping rainfall terrain",
                    "natural hazard risk mapping machine learning GIS",
                ],
                max_journals=5,
            )

        journal = result["candidates"][0]
        self.assertEqual(journal["name"], "Natural Hazards")
        self.assertEqual(journal["similar_works_count"], 3)
        self.assertEqual(journal["query_coverage"], 2)
        self.assertEqual(len(journal["publication_precedents"]), 3)
        self.assertGreater(journal["literature_evidence_score"], 0)
        self.assertEqual(journal["source_recent_works_count"], 3000)
        self.assertEqual(journal["similarity_rate_per_10k"], 10.0)

    def test_openalex_similar_works_exclude_the_target_paper_itself(self):
        payload = {
            "results": [
                {
                    "id": "https://openalex.org/W-TARGET",
                    "display_name": "Exact Blind Test Paper Title",
                    "publication_year": 2025,
                    "type": "article",
                    "primary_location": {
                        "source": {
                            "id": "https://openalex.org/S1",
                            "display_name": "Target Journal",
                            "type": "journal",
                            "issn_l": "1234-5678",
                            "issn": ["1234-5678"],
                        }
                    },
                    "primary_topic": {"display_name": "Target Topic"},
                },
                {
                    "id": "https://openalex.org/W-RELATED",
                    "display_name": "A genuinely related paper",
                    "publication_year": 2024,
                    "type": "article",
                    "primary_location": {
                        "source": {
                            "id": "https://openalex.org/S1",
                            "display_name": "Target Journal",
                            "type": "journal",
                            "issn_l": "1234-5678",
                            "issn": ["1234-5678"],
                        }
                    },
                    "primary_topic": {"display_name": "Target Topic"},
                },
            ]
        }
        response = unittest.mock.Mock(status_code=200, json=lambda: payload)
        response.raise_for_status.return_value = None

        with patch.dict(os.environ, {"OPENALEX_API_KEY": "test-key"}), \
            patch("scripts.similar_works.requests.get", return_value=response):
            result = discover_journal_candidates(
                ["blind test topic query"],
                excluded_titles=["Exact Blind Test Paper Title"],
                excluded_work_ids=["https://openalex.org/W-TARGET"],
            )

        journal = result["candidates"][0]
        self.assertEqual(journal["similar_works_count"], 1)
        self.assertEqual(journal["publication_precedents"][0]["openalex_id"], "https://openalex.org/W-RELATED")

    def test_literature_precedent_outranks_unverified_high_if_journal(self):
        profile = merge_paper_profile(
            infer_paper_profile("flash flood susceptibility mapping rainfall terrain"),
            {
                "direction_summary": "flash-flood susceptibility and natural-hazard mapping",
                "research_object": "flash floods",
                "search_queries": ["flash flood susceptibility mapping rainfall terrain"],
            },
        )
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "High Impact AI Letters",
                    "impact_factor": "18",
                    "jcr_quartile": "Q1",
                    "sci_type": "SCIE",
                    "field": "Artificial Intelligence",
                    "_sources": ["journal-index", "openalex"],
                },
                {
                    "name": "Natural Hazards",
                    "impact_factor": "3.7",
                    "jcr_quartile": "Q2",
                    "sci_type": "SCIE",
                    "field": "Natural Hazards",
                    "similar_works_count": 3,
                    "query_coverage": 2,
                    "literature_evidence_score": 18,
                    "publication_precedents": [
                        {"title": "Flash-flood susceptibility mapping", "year": 2024},
                        {"title": "Rainfall controls on flash floods", "year": 2023},
                    ],
                    "_sources": ["journal-index", "openalex-works"],
                },
            ],
        )

        self.assertEqual(ranked[0]["name"], "Natural Hazards")
        self.assertEqual(ranked[0]["fit_confidence"], "强")
        self.assertNotEqual(ranked[1]["tier"], "推荐")

    def test_scope_specific_journal_name_outranks_broad_high_if_with_equal_precedents(self):
        profile = merge_paper_profile(
            infer_paper_profile("hybrid energy storage for renewable grid integration"),
            {
                "direction_summary": "hybrid energy storage for renewable power-grid integration",
                "primary_field": "power and energy systems",
                "specialty": "hybrid energy storage and grid integration",
                "research_object": "grid-connected hybrid energy storage systems",
                "target_audience": ["power-system engineers", "energy-storage researchers"],
            },
        )
        common = {
            "sci_type": "SCIE",
            "similar_works_count": 4,
            "query_coverage": 3,
            "literature_evidence_score": 24,
            "_sources": ["journal-index", "openalex-works"],
        }
        ranked = rank_metric_records(
            profile,
            [
                {
                    **common,
                    "name": "Environmental Chemistry Letters",
                    "impact_factor": "15",
                    "jcr_quartile": "Q1",
                    "field": "Environmental Chemistry",
                },
                {
                    **common,
                    "name": "Journal of Energy Storage",
                    "impact_factor": "9",
                    "jcr_quartile": "Q1",
                    "field": "Energy Storage",
                },
            ],
        )

        self.assertEqual(ranked[0]["name"], "Journal of Energy Storage")
        self.assertIn("期刊名称覆盖核心词", "；".join(ranked[0]["fit_reasons"]))

    def test_candidate_topics_survive_local_partition_enrichment(self):
        metrics_record = {
            "name": "Journal of Energy Storage",
            "cas_partition_2025": "2区",
            "_sources": ["journal-index"],
        }
        candidate = {
            "name": "Journal of Energy Storage",
            "field": "Hybrid Energy Storage; Renewable Grid Integration",
            "openalex_topics": ["Hybrid Energy Storage", "Renewable Grid Integration"],
            "_sources": ["openalex-works"],
        }

        selector._merge_candidate_evidence(metrics_record, candidate)

        self.assertEqual(
            metrics_record["field"],
            "Hybrid Energy Storage; Renewable Grid Integration",
        )
        self.assertIn("openalex-works", metrics_record["_sources"])

    def test_positioning_input_does_not_create_submission_role(self):
        profile = merge_paper_profile(
            infer_paper_profile("clinical NLP benchmark"),
            {
                "positioning": {
                    "level": "strong",
                    "confidence": "high",
                    "evidence": ["five tasks", "three model families"],
                }
            },
        )
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "High Digital Medicine Journal",
                    "impact_factor": "15",
                    "jcr_quartile": "Q1",
                    "sci_type": "SCIE",
                    "field": "clinical NLP medicine",
                    "similar_works_count": 3,
                    "query_coverage": 2,
                    "literature_evidence_score": 18,
                    "_sources": ["journal-index", "openalex-works"],
                }
            ],
        )

        banded = assign_submission_bands(ranked, profile=profile)

        self.assertNotIn("positioning", profile)
        self.assertEqual(banded[0]["journal_level"], "高位")
        self.assertEqual(banded[0]["submission_band"], "高位候选")

    def test_openalex_retries_transient_rate_limit(self):
        rate_limited = unittest.mock.Mock(status_code=429, headers={"Retry-After": "0"})
        rate_limited.raise_for_status.side_effect = requests.HTTPError("429")
        ok = unittest.mock.Mock(
            status_code=200,
            headers={},
            json=lambda: {"results": []},
        )
        ok.raise_for_status.return_value = None

        with patch.dict(os.environ, {"OPENALEX_API_KEY": "test-key"}), \
            patch("scripts.similar_works.requests.get", side_effect=[rate_limited, ok]) as get, \
            patch("scripts.similar_works.time.sleep") as sleep:
            result = discover_journal_candidates(["rate limited query"])

        self.assertEqual(result["errors"], [])
        self.assertEqual(get.call_count, 2)
        sleep.assert_called_once()

    def test_report_states_candidate_discovery_boundary(self):
        profile = merge_paper_profile(
            infer_paper_profile("groundwater nitrate hydrochemistry"),
            {
                "direction_summary": "groundwater nitrate hydrochemistry",
                "research_object": "groundwater nitrate",
                "positioning": {"level": "unknown", "confidence": "low", "evidence": []},
            },
        )
        ranked = rank_metric_records(
            profile,
            [
                {
                    "name": "Water Research",
                    "impact_factor": "12.8",
                    "cas_partition_2025": "1区",
                    "jcr_quartile": "Q1",
                    "sci_type": "SCIE",
                    "field": "Water Resources",
                    "similar_works_count": 2,
                    "query_coverage": 2,
                    "literature_evidence_score": 18,
                    "_sources": ["journal-index", "openalex-works"],
                }
            ],
        )
        banded = assign_submission_bands(ranked, profile=profile)
        report = format_selection_report(profile, banded)

        self.assertEqual(banded[0]["journal_level"], "高位")
        self.assertEqual(banded[0]["submission_band"], "高位候选")
        self.assertIn("不声称存在唯一最优期刊", report)
        self.assertIn("官网 scope 待核验", report)
        self.assertNotIn("摘要可见定位", report)
        self.assertNotIn("主投", report)
        self.assertNotIn("保底", report)

    def test_benchmark_manifest_has_sixty_unique_cross_disciplinary_cases(self):
        manifest = json.loads(
            Path("benchmarks/corpus_manifest.json").read_text(encoding="utf-8")
        )
        strata = {}
        for row in manifest:
            strata[row["stratum"]] = strata.get(row["stratum"], 0) + 1

        self.assertEqual(len(manifest), 60)
        self.assertEqual(len({row["doi"] for row in manifest}), 60)
        self.assertEqual(len(strata), 20)
        self.assertEqual(set(strata.values()), {3})

    def test_expert_benchmark_scores_relevance_not_only_original_journal(self):
        predictions = [
            {
                "case_id": "B001",
                "candidates": [
                    {"journal": "Broad Journal"},
                    {"journal": "Specialist Journal"},
                    {"journal": "Wrong Journal"},
                ],
            }
        ]
        labels = []
        for rater in ("r1", "r2"):
            labels.extend(
                [
                    {
                        "case_id": "B001", "candidate_journal": "Broad Journal",
                        "rater_id": rater, "fit_label": "borderline",
                        "community_match": "yes", "journal_scope_type": "generalist",
                        "scope_checked": "yes",
                    },
                    {
                        "case_id": "B001", "candidate_journal": "Specialist Journal",
                        "rater_id": rater, "fit_label": "suitable",
                        "community_match": "yes", "journal_scope_type": "specialist",
                        "scope_checked": "yes",
                    },
                    {
                        "case_id": "B001", "candidate_journal": "Wrong Journal",
                        "rater_id": rater, "fit_label": "unsuitable",
                        "community_match": "no", "journal_scope_type": "specialist",
                        "scope_checked": "no",
                    },
                ]
            )

        scored = score_benchmark(predictions, labels)
        auxiliary = auxiliary_exact_journal_metrics(
            predictions,
            [{"case_id": "B001", "original_journal": "Specialist Journal"}],
        )

        self.assertEqual(scored["summary"]["acceptable_recall@3"], 1.0)
        self.assertTrue(scored["summary"]["ready"])
        self.assertEqual(scored["summary"]["community_recall@3"], 1.0)
        self.assertGreater(scored["summary"]["ndcg@3"], 0)
        self.assertAlmostEqual(scored["summary"]["generalist_exposure@3"], 1 / 3, places=4)
        self.assertEqual(auxiliary["exact_original_journal_hit@3"], 1)
        self.assertIn("Auxiliary only", auxiliary["note"])

    def test_volume_normalization_favors_dense_specialist_evidence(self):
        def candidate(recent_works):
            entry = similar_works._new_candidate(
                {
                    "id": f"https://openalex.org/S{recent_works}",
                    "display_name": f"Journal {recent_works}",
                    "type": "journal",
                }
            )
            entry["publication_precedents"] = [{"title": str(i)} for i in range(3)]
            entry["_query_indexes"] = {0, 1}
            entry["_rank_signal"] = 1.0
            entry["_best_rank"] = 1
            entry["source_recent_works_count"] = recent_works
            return similar_works._finish_candidate(entry)

        broad = candidate(100000)
        specialist = candidate(1000)

        self.assertGreater(specialist["similarity_rate_per_10k"], broad["similarity_rate_per_10k"])
        self.assertGreater(specialist["literature_evidence_score"], broad["literature_evidence_score"])

    def test_high_impact_factor_does_not_change_scope_candidate_status(self):
        profile = {
            "primary_field": "hydrogeology",
            "specialty": "groundwater nitrate",
            "research_object": "groundwater nitrate",
            "categories": [],
        }
        common = {
            "field": "hydrogeology groundwater nitrate",
            "similar_works_count": 3,
            "query_coverage": 2,
            "literature_evidence_score": 18,
            "sci_type": "SCIE",
            "_sources": ["openalex-works"],
        }
        ranked = rank_metric_records(
            profile,
            [
                {**common, "name": "High Standing Journal", "impact_factor": "25", "jcr_quartile": "Q1"},
                {**common, "name": "Regular Standing Journal", "impact_factor": "2", "jcr_quartile": "Q3"},
                {**common, "name": "Fourth Quartile Journal", "impact_factor": "1", "xinrui_partition_2026": "4区"},
            ],
        )

        self.assertEqual(len({row["tier"] for row in ranked}), 1)
        self.assertNotEqual(ranked[0]["journal_level"], ranked[1]["journal_level"])

    def test_missing_openalex_key_is_an_explicit_retrieval_degradation(self):
        with patch.dict(os.environ, {"OPENALEX_API_KEY": ""}, clear=False), \
            patch("scripts.similar_works.requests.get") as get:
            result = discover_journal_candidates(["groundwater nitrate"])

        get.assert_not_called()
        self.assertEqual(result["candidates"], [])
        self.assertIn("API key未配置", result["errors"][0])

    def test_specialist_title_channel_recalls_domain_journal_without_blacklist(self):
        candidates = search_index_journals(["energy", "storage"], limit=20)

        self.assertIn(
            "journal of energy storage",
            {row["name"].lower() for row in candidates},
        )
        self.assertTrue(all("journal-title-index" in row["_sources"] for row in candidates))

    def test_official_scope_evidence_requires_official_url_and_real_text(self):
        profile = {
            "primary_field": "hydrogeology",
            "specialty": "groundwater nitrate reference conditions",
            "research_object": "groundwater nitrate",
            "research_question": "reference condition estimation",
            "contribution_type": "hydrochemical study",
            "target_audience": ["hydrogeologists"],
            "topic_evidence": ["aquifer", "nitrate"],
        }
        text = (
            "This journal publishes research on groundwater, aquifers, hydrogeology, "
            "groundwater quality, nitrate transport, hydrochemical processes, and reference "
            "conditions. It welcomes field studies and methodological advances that improve "
            "understanding of subsurface water resources and contaminant behavior."
        )

        evidence = verify_official_scope(
            profile,
            "Hydrogeology Journal",
            text,
            "https://link.springer.com/journal/10040/aims-and-scope",
            publisher_domain_confirmed=True,
        )
        queue = build_scope_verification_queue(profile, [{"name": "Hydrogeology Journal"}])

        self.assertTrue(evidence["scope_checked"])
        self.assertTrue(evidence["scope_verified"])
        self.assertEqual(queue[0]["status"], "待核验")
        verified_queue = build_scope_verification_queue(profile, [evidence])
        self.assertEqual(verified_queue[0]["status"], "已核验")
        self.assertEqual(verified_queue[0]["source_url"], evidence["scope_source_url"])
        with self.assertRaises(ValueError):
            verify_official_scope(
                profile,
                "Hydrogeology Journal",
                text,
                "https://link.springer.com/journal/10040/aims-and-scope",
            )
        with self.assertRaises(ValueError):
            verify_official_scope(profile, "Hydrogeology Journal", text, "https://openalex.org/S1")

        manual_record = {
            **evidence,
            "scope_source_domain_confirmed": False,
        }
        candidate = {"name": "Hydrogeology Journal"}
        from scripts.scope_evidence import apply_scope_record
        apply_scope_record(candidate, [manual_record])
        self.assertNotIn("scope_verified", candidate)

    def test_cross_language_scope_check_is_unresolved_not_incompatible(self):
        profile = {
            "primary_field": "水文地质学",
            "specialty": "地下水硝酸盐",
            "research_object": "浅层含水层",
            "research_question": "参考条件估计",
            "contribution_type": "水化学研究",
            "target_audience": ["地下水研究者"],
        }
        evidence = verify_official_scope(
            profile,
            "Hydrogeology Journal",
            (
                "This journal publishes original research on groundwater systems, aquifers, "
                "water quality, contaminant transport, and field hydrogeology. It welcomes "
                "methodological and applied studies that improve understanding of subsurface "
                "water resources and related environmental processes."
            ),
            "https://link.springer.com/journal/10040/aims-and-scope",
            publisher_domain_confirmed=True,
        )

        self.assertEqual(evidence["scope_match"], "unresolved")
        self.assertFalse(evidence["scope_verified"])

    def test_scope_evidence_reranks_the_same_candidate_pool(self):
        profile = {
            "primary_field": "hydrogeology",
            "specialty": "groundwater nitrate",
            "research_object": "groundwater nitrate",
            "research_question": "nitrate reference conditions",
            "contribution_type": "field study",
            "target_audience": ["hydrogeologists"],
            "categories": [],
        }
        scope = verify_official_scope(
            profile,
            "Specialist Groundwater Journal",
            (
                "The journal publishes hydrogeology, groundwater nitrate, aquifer quality, "
                "reference condition, contaminant transport, and field studies for groundwater "
                "researchers. It welcomes original field investigations and applied methods for "
                "understanding subsurface water-quality processes."
            ),
            "https://link.springer.com/journal/example/aims-and-scope",
            publisher_domain_confirmed=True,
        )
        common = {
            "sci_type": "SCIE",
            "similar_works_count": 2,
            "query_coverage": 1,
            "literature_evidence_score": 10,
            "_sources": ["openalex-works"],
        }
        ranked = selector.rerank_with_scope_evidence(
            profile,
            [
                {**common, "name": "Broad Science Journal", "field": "general science"},
                {**common, "name": "Specialist Groundwater Journal", "field": "groundwater"},
            ],
            [scope],
        )

        self.assertEqual(ranked[0]["name"], "Specialist Groundwater Journal")
        self.assertTrue(ranked[0]["scope_verified"])
        self.assertGreater(ranked[0]["fit_score"], ranked[1]["fit_score"])

    def test_profile_protocol_reports_cross_model_disagreement(self):
        base = {
            "direction_summary": "groundwater nitrate hydrogeology",
            "primary_field": "hydrogeology",
            "specialty": "groundwater nitrate",
            "research_object": "shallow aquifer nitrate",
            "research_question": "reference concentration estimation",
            "contribution_type": "applied hydrochemical framework",
            "target_audience": ["hydrogeologists"],
            "methods": ["hydrochemistry"],
            "exclusions": ["artificial intelligence"],
            "search_queries": ["groundwater nitrate aquifer", "nitrate reference hydrochemistry"],
        }
        divergent = {
            **base,
            "primary_field": "artificial intelligence",
            "specialty": "machine learning algorithms",
            "research_object": "neural network architecture",
            "research_question": "classification benchmark",
        }

        self.assertTrue(validate_profile(base)["valid"])
        comparison = compare_profiles([base, divergent])
        self.assertIn(comparison["status"], {"low", "medium"})
        self.assertIn("primary_field", comparison["disagreements"])

        with patch.object(selector, "discover_journal_candidates", return_value={"queries": [], "candidates": [], "errors": []}), \
            patch.object(selector, "discover_specialist_candidates", return_value=[]), \
            patch.object(selector, "search_candidates", return_value=[]):
            bundle = selector.select_journals(
                "cross-disciplinary abstract",
                independent_profiles=[base, divergent],
                max_candidates=3,
                request_delay=0,
            )

        self.assertEqual(bundle["profile"]["profile_consistency"]["status"], comparison["status"])
        self.assertGreaterEqual(len(bundle["profile"]["retrieval_search_queries"]), 2)
        self.assertTrue(bundle["profile"]["profile_validation"]["valid"])

    def test_profile_consistency_tokenizes_related_chinese_phrases(self):
        first = {
            "direction_summary": "地下水硝酸盐水文地球化学",
            "primary_field": "水文地质学",
            "specialty": "地下水硝酸盐污染",
            "research_object": "浅层地下水硝酸盐",
            "research_question": "参考条件如何估计",
            "contribution_type": "水文地球化学研究",
            "target_audience": ["水文地质研究者"],
            "methods": ["水化学"],
            "exclusions": ["人工智能算法"],
            "search_queries": ["地下水 硝酸盐 参考条件", "含水层 水化学 硝酸盐"],
        }
        second = {
            **first,
            "primary_field": "地下水水文地质",
            "specialty": "浅层地下水硝酸盐污染",
            "research_object": "含水层硝酸盐",
        }

        comparison = compare_profiles([first, second])

        self.assertNotEqual(comparison["status"], "low")

    def test_incomplete_expert_labels_block_benchmark_release(self):
        predictions = [
            {"case_id": "B001", "candidates": [{"journal": "Candidate Journal"}]}
        ]
        labels = [
            {
                "case_id": "B001",
                "candidate_journal": "Candidate Journal",
                "rater_id": "r1",
                "fit_label": "suitable",
                "community_match": "yes",
                "journal_scope_type": "specialist",
                "scope_checked": "yes",
            }
        ]

        scored = score_benchmark(predictions, labels, min_raters=2)

        self.assertFalse(scored["summary"]["ready"])
        self.assertEqual(scored["summary"]["incomplete_candidate_labels"], 1)
        self.assertEqual(scored["summary"]["label_coverage"], 0.0)

    def test_unfilled_expert_nomination_blocks_benchmark_release(self):
        predictions = [
            {"case_id": "B001", "candidates": [{"journal": "System Journal"}]}
        ]
        labels = []
        for rater in ("r1", "r2"):
            labels.append(
                {
                    "case_id": "B001",
                    "candidate_journal": "System Journal",
                    "rater_id": rater,
                    "fit_label": "suitable",
                    "community_match": "yes",
                    "journal_scope_type": "specialist",
                    "scope_checked": "yes",
                }
            )
        labels.append(
            {
                "case_id": "B001",
                "candidate_journal": "Expert Nominated Journal",
                "rater_id": "r1",
                "fit_label": "",
                "community_match": "",
                "journal_scope_type": "",
                "scope_checked": "",
            }
        )

        scored = score_benchmark(predictions, labels, min_raters=2)

        self.assertFalse(scored["summary"]["ready"])
        self.assertEqual(scored["summary"]["label_coverage"], 0.5)

    def test_benchmark_runner_never_reads_or_requires_original_journal_truth(self):
        fake_bundle = {
            "profile": {"direction_summary": "test profile"},
            "retrieval": {"queries": ["test query"]},
            "results": [
                {
                    "name": "Candidate Journal",
                    "tier": "备选",
                    "fit_confidence": "中",
                    "journal_level": "中位",
                    "fit_score": 18,
                }
            ],
        }
        with patch("scripts.benchmark_run.select_journals", return_value=fake_bundle) as select:
            predictions = run_benchmark(
                [
                    {
                        "case_id": "B001",
                        "title": "Blind paper title",
                        "abstract": "Blind paper abstract",
                        "keywords": ["blind", "test"],
                        "stratum": "test",
                    }
                ],
                request_delay=0,
            )

        self.assertEqual(predictions[0]["candidates"][0]["journal"], "Candidate Journal")
        self.assertNotIn("original_journal", predictions[0])
        self.assertIn("Blind paper title", select.call_args.kwargs["excluded_titles"])

    def test_annotation_pool_deduplicates_systems_and_accepts_expert_nominations(self):
        predictions = [
            {"case_id": "B001", "candidates": [{"journal": "Journal of Testing"}]},
            {"case_id": "B001", "candidates": [{"journal": "JOURNAL OF TESTING"}]},
        ]
        nominations = [{"case_id": "B001", "journal": "Specialist Test Journal"}]
        with tempfile.TemporaryDirectory() as tmpdir:
            output = os.path.join(tmpdir, "labels.csv")
            count = build_annotation_sheet(
                predictions,
                output,
                rater_ids=("r1", "r2"),
                nominations=nominations,
            )
            with open(output, "r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))

        self.assertEqual(count, 4)
        self.assertEqual(len(rows), 4)
        self.assertNotIn("candidate_rank", rows[0])
        self.assertEqual(
            {row["candidate_journal"].lower() for row in rows},
            {"journal of testing", "specialist test journal"},
        )

    def test_bundled_index_passes_release_gate_and_missing_provenance_fails(self):
        bundled = audit_index("assets/sci_select_journals.sqlite")
        self.assertTrue(bundled["passed"])

        with tempfile.TemporaryDirectory() as tmpdir:
            database = os.path.join(tmpdir, "bad.sqlite")
            write_sqlite_index(
                {
                    "meta": {"schema": "sci-select-journal-index-v1"},
                    "journals": [{"title": "Example Journal", "cas_2025": "2区"}],
                },
                database,
            )
            failed = audit_index(database)

        self.assertFalse(failed["passed"])
        self.assertEqual(failed["summary"]["provenance_errors"], 1)

    def test_official_finder_checklist_is_manual_and_opt_in(self):
        checklist = build_finder_checklist(
            title="Status separation before estimating nitrate reference conditions",
            abstract="This study tested a geology-constrained hydrochemical framework.",
            keywords=["groundwater", "nitrate", "hydrochemistry"],
        )
        report = format_finder_checklist(checklist)

        self.assertIn("Elsevier Journal Finder", report)
        self.assertIn("Springer Nature Journal Finder", report)
        self.assertIn("groundwater; nitrate; hydrochemistry", report)
        self.assertIn("手动打开", report)
        self.assertIn("可选", report)
        self.assertNotIn("cookie", report.lower())
        self.assertNotIn("自动登录", report)
        self.assertNotIn("模拟登录", report)


if __name__ == "__main__":
    unittest.main()
